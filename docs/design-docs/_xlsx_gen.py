#!/usr/bin/env python3
"""XLSX generator using openpyxl with OperaMind-style formatting.

Style reference (matches OperaMind generate_design_docs.py):
  - Headers: dark navy #1E293B bg, white bold Yu Gothic 11pt
  - Titles:  navy #1E293B bold Yu Gothic 16pt
  - Labels:  light gray #F1F5F9 bg, slate bold Yu Gothic 10pt
  - Values:  navy #1E293B Yu Gothic 10pt
  - Sections: light blue #DBEAFE bg, indigo bold Yu Gothic 11pt
  - Borders: thin all sides, wrap text + top align
"""

from openpyxl import Workbook as OpxlWorkbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

DATE = datetime.now().strftime("%Y/%m/%d")
VERSION = "1.0"

# ── OperaMind-style constants ──────────────────────────────────
HEADER_FILL  = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT  = Font(name="游ゴシック", bold=True, size=11, color="FFFFFF")
TITLE_FONT   = Font(name="游ゴシック", bold=True, size=16, color="1E293B")
LABEL_FILL   = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
LABEL_FONT   = Font(name="游ゴシック", bold=True, size=10, color="475569")
VALUE_FONT   = Font(name="游ゴシック", size=10, color="1E293B")
SECTION_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
SECTION_FONT = Font(name="游ゴシック", bold=True, size=11, color="1E40AF")
GREEN_FILL   = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP   = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")

# ── Style IDs (backward-compatible) ────────────────────────────
STYLE_HEADER     = 0   # bold white on dark navy, centered
STYLE_NORMAL     = 1   # value font, thin border, wrap
STYLE_TITLE      = 2   # bold 16pt navy, for title row
STYLE_DOCINFO_K  = 3   # bold label on light gray
STYLE_SECTION    = 4   # bold indigo on light blue
STYLE_BOLD       = 5   # bold + border, no fill

# ── XlsxWriter ─────────────────────────────────────────────────
class XlsxWriter:
    """Collect sheets as list-of-rows data, then render via openpyxl on save()."""

    def __init__(self):
        self._wb = OpxlWorkbook()
        self._wb.remove(self._wb.active)
        self._sheets = {}       # name → list of rows
        self._col_widths = {}   # name → list of int
        self._sheet_order = []

    def add_sheet(self, name, data, col_widths=None):
        """Register a sheet.  *data* is a list of rows; each row is a list of
        ``(text, style_id)`` tuples (or bare str, defaulting to STYLE_NORMAL).
        A row may also be a ``(list_of_tuples, merge_range_str)`` pair.  """
        self._sheet_order.append(name)
        self._sheets[name] = data
        if col_widths:
            self._col_widths[name] = col_widths

    # ── internal helpers ───────────────────────────────────

    @staticmethod
    def _apply(cell, style_id):
        if style_id == STYLE_HEADER:
            cell.font = HEADER_FONT; cell.fill = HEADER_FILL
            cell.alignment = CENTER; cell.border = THIN
        elif style_id == STYLE_TITLE:
            cell.font = TITLE_FONT; cell.border = THIN
        elif style_id == STYLE_DOCINFO_K:
            cell.font = LABEL_FONT; cell.fill = LABEL_FILL; cell.border = THIN
        elif style_id == STYLE_SECTION:
            cell.font = SECTION_FONT; cell.fill = SECTION_FILL; cell.border = THIN
        elif style_id == STYLE_BOLD:
            cell.font = Font(name="游ゴシック", bold=True, size=10, color="1E293B")
            cell.border = THIN
        else:   # STYLE_NORMAL
            cell.font = VALUE_FONT; cell.alignment = WRAP; cell.border = THIN

    def save(self, path):
        """Render all sheets into an .xlsx file."""
        for sheet_name in self._sheet_order:
            data = self._sheets[sheet_name]
            ws = self._wb.create_sheet(title=sheet_name)

            if sheet_name in self._col_widths:
                for i, w in enumerate(self._col_widths[sheet_name], 1):
                    ws.column_dimensions[get_column_letter(i)].width = w

            if not data:
                continue

            merges = []
            for ri, row_data in enumerate(data, 1):
                merge_range = None
                if isinstance(row_data, tuple):
                    row_data, merge_range = row_data

                for ci, cell_data in enumerate(row_data, 1):
                    if isinstance(cell_data, tuple):
                        text, style = cell_data
                    else:
                        text, style = str(cell_data) if cell_data else "", STYLE_NORMAL
                    cell = ws.cell(row=ri, column=ci, value=str(text) if text else "")
                    XlsxWriter._apply(cell, style)

                if merge_range:
                    # merge_range strings from T() use row 1 → fix to actual row
                    merges.append(_fix_merge_row(merge_range, ri))
                else:
                    # auto-merge: if leading cell is TITLE or SECTION,
                    # merge all consecutive cells with the same style
                    if row_data and isinstance(row_data[0], tuple):
                        first_style = row_data[0][1]
                        if first_style in (STYLE_TITLE, STYLE_SECTION):
                            end = 1
                            for cd in row_data[1:]:
                                if isinstance(cd, tuple) and cd[1] == first_style:
                                    end += 1
                                else:
                                    break
                            if end > 1:
                                sc = get_column_letter(1)
                                ec = get_column_letter(end)
                                merges.append(f"{sc}{ri}:{ec}{ri}")

                # row heights
                if row_data and isinstance(row_data[0], tuple):
                    s = row_data[0][1]
                    if s == STYLE_TITLE:
                        ws.row_dimensions[ri].height = 36
                    elif s == STYLE_SECTION:
                        ws.row_dimensions[ri].height = 22

            for m in merges:
                ws.merge_cells(m)

        self._wb.save(path)
        print(f"  -> {os.path.basename(path)}")


def _fix_merge_row(m, row):
    """Replace row numbers in merge-range string with *row*."""
    import re
    return re.sub(r'\d+', str(row), m)


# ── Row helpers (backward-compatible names) ────────────────────

def H(items):
    """Header row: dark navy bg, white bold, centered."""
    return [(str(x) if x else "", STYLE_HEADER) for x in items]

def R(items):
    """Normal data row: thin border, value font, wrap."""
    return [(str(x) if x else "", STYLE_NORMAL) for x in items]

def K(items):
    """Key-value row: first cell label-style, rest normal."""
    out = [(str(items[0]) if items[0] else "", STYLE_DOCINFO_K)]
    for x in items[1:]:
        out.append((str(x) if x else "", STYLE_NORMAL))
    return out

def T(text, cols=8):
    """Title row: bold 16pt, auto-merged across *cols*."""
    row = [(text, STYLE_TITLE)] + [("", STYLE_TITLE)] * (cols - 1)
    sc = get_column_letter(1)
    ec = get_column_letter(cols)
    return (row, f"{sc}1:{ec}1")

def S(text, cols=8):
    """Section header: bold indigo on light-blue bg."""
    return [(text, STYLE_SECTION)] + [("", STYLE_SECTION)] * (cols - 1)

def E(n=8):
    """Empty row (thin border on every cell)."""
    return [("", STYLE_NORMAL)] * n

# Aliases used by older gen scripts
hdr_row = H
empty_row = E

# ── Standard blocks ────────────────────────────────────────────

def doc_info_8():
    """8-column document metadata block (OperaMind-style)."""
    return [
        S("文書管理情報", 8),
        K(["システムID", "STRUTSLAB-001", "", "", "システム名",
           "電力設備巡視点検管理システム", "", ""]),
        K(["サブシステム名", "", "", "", "開発言語",
           "Java 1.8 / Struts1 / MyBatis", "", ""]),
        K(["データベース", "H2 Database（ファイルモード）", "", "",
           "文書番号", "", "", ""]),
        K(["版数", VERSION, "", "", "作成日", DATE, "", ""]),
        K(["作成者", "システム開発部", "", "", "承認者", "", "", ""]),
    ]

# legacy uppercase alias
DOC_INFO_8 = doc_info_8

def doc_info_6():
    """6-column document metadata block (OperaMind-style)."""
    return [
        S("文書管理情報", 6),
        K(["システムID", "STRUTSLAB-001", "", "", "サブシステム名", ""]),
        K(["画面ID", "", "", "", "画面名", ""]),
        K(["プログラムID", "", "", "", "文書番号", ""]),
        K(["版数", VERSION, "", "", "作成日", DATE]),
        K(["作成者", "システム開発部", "", "", "区分", "新規"]),
    ]

REVISION_HISTORY = [
    T("改訂履歷", 7),
    H(["版数", "改訂日", "改訂内容", "改訂者", "承認者", "区分", "備考"]),
    R([VERSION, DATE, "初版作成", "システム開発部", "", "新規", ""]),
]

def make_standard_workbook(main_sheet_name, main_data, main_col_widths=None, **extra_sheets):
    """Create workbook with 改訂履歷 + main sheet + optional extras."""
    w = XlsxWriter()
    w.add_sheet("改訂履歷", REVISION_HISTORY, col_widths=[8, 14, 40, 18, 18, 10, 24])
    w.add_sheet(main_sheet_name, main_data, col_widths=main_col_widths)
    for sname, sdata in extra_sheets.items():
        w.add_sheet(sname, sdata)
    return w
